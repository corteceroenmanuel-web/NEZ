import os
from datetime import datetime
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash 
import sqlite3 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import hashlib
import platform
import uuid
from datetime import datetime, timedelta
from functools import wraps

CLAVE_SECRETA = "PLA$r0l4y_N3Z*"

def obtener_hwid():
    raw = f"{platform.node()}-{uuid.getnode()}-{platform.processor()}"
    return hashlib.sha256(raw.encode()).hexdigest()[:12].upper()

def generar_serial(hwid, fecha_limite_str="VITALICIA"):
    cadena = f"{hwid}-{fecha_limite_str}-{CLAVE_SECRETA}"
    hash_code = hashlib.sha256(cadena.encode()).hexdigest()[:16].upper()
    return f"{fecha_limite_str}-{hash_code}"

def verificar_licencia_actual():
    config = Configuracion.query.first()
    if not config or not config.licencia_key:
        return False, "Sin licencia registrada.", None

    hwid_actual = obtener_hwid()
    key = config.licencia_key

    if key.startswith("VITALICIA-"):
        serial_esperado = generar_serial(hwid_actual, "VITALICIA")
        if key == serial_esperado:
            return True, "Licencia Vitalicia Activa", None
        return False, "La licencia no pertenece a este equipo.", None

    try:
        partes = key.split("-")
        fecha_str = f"{partes[0]}-{partes[1]}-{partes[2]}"
        serial_esperado = generar_serial(hwid_actual, fecha_str)

        if key != serial_esperado:
            return False, "La licencia no es válida para este equipo.", None

        fecha_limite = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        if datetime.now().date() > fecha_limite:
            return False, f"Licencia vencida el {fecha_str}.", fecha_limite

        return True, f"Licencia Activa hasta {fecha_str}", fecha_limite
    except Exception:
        return False, "Formato de licencia inválido.", None
    
def licencia_requerida(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        valida, mensaje, fecha = verificar_licencia_actual()
        if not valida:
            flash(f"🔒 {mensaje} Por favor activa una licencia para continuar.", "warning")
            return redirect(url_for('activar_licencia'))
        return f(*args, **kwargs)
    return decorated_function

app = Flask(__name__)
# Clave secreta para encriptar las sesiones del navegador de forma segura
app.secret_key = "clave_secreta_rolay_negocio"

# CONFIGURACIÓN DE LA BASE DE DATOS (SQLite local)
base_dir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(base_dir, 'negocio.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ========================================================
# MODELOS DE LA BASE DE DATOS (ESTRUCTURA DE LAS TABLAS)
# ========================================================

class Venta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'))
    total = db.Column(db.Float)
    es_credito = db.Column(db.Boolean, default=False)
    # Relación con los detalles:
    detalles = db.relationship('DetalleVenta', backref='venta', lazy=True)

class DetalleVenta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venta_id = db.Column(db.Integer, db.ForeignKey('venta.id'))
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'))
    cantidad = db.Column(db.Integer)
    precio_unitario = db.Column(db.Float)
    
    # Relación para acceder directamente al producto
    producto = db.relationship('Producto')

class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    stock = db.Column(db.Integer, default=0)
    costo_compra = db.Column(db.Float, default=0.0)
    precio_venta = db.Column(db.Float, default=0.0)
    proveedor_id = db.Column(db.Integer, db.ForeignKey('proveedor.id'), nullable=True)
    ventas_totales = db.Column(db.Integer, default=0)

    # Relación para poder acceder a producto.proveedor.nombre directamente
    proveedor = db.relationship('Proveedor', backref='productos')

    # 🛡️ INIT SEGURO CORREGIDO
    def __init__(self, **kwargs):
        # 1. Validar Nombre
        nombre_val = kwargs.get('nombre')
        if not nombre_val or not str(nombre_val).strip():
            kwargs['nombre'] = f"Producto Nuevo {datetime.now().strftime('%H%M%S')}"
        else:
            kwargs['nombre'] = str(nombre_val).strip()

        # 2. Validar y asegurar el Proveedor ID
        prov_id = kwargs.get('proveedor_id')
        if prov_id is not None and str(prov_id).strip() != "":
            try:
                kwargs['proveedor_id'] = int(prov_id)
            except ValueError:
                kwargs['proveedor_id'] = None
        else:
            kwargs['proveedor_id'] = None

        # 3. Pasar TODOS los argumentos (incluyendo proveedor_id, stock, costo, etc.) a SQLAlchemy
        super(Producto, self).__init__(**kwargs)

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    telefono = db.Column(db.String(50))
    deuda_total = db.Column(db.Float, default=0.0)
    
    # Esta relación permite acceder a c.deudas desde el HTML:
    deudas = db.relationship('Deuda', backref='cliente', lazy=True)

class Deuda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    tipo = db.Column(db.String(50), nullable=False) 
    monto_inicial = db.Column(db.Float, nullable=False)
    saldo_pendiente = db.Column(db.Float, nullable=False)
    descripcion = db.Column(db.Text, nullable=True) 
    fecha = db.Column(db.DateTime, default=db.func.now())
    estado = db.Column(db.String(20), default="Pendiente") 
    
    abonos = db.relationship('HistorialAbono', backref='deuda', lazy=True)

class HistorialAbono(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    deuda_id = db.Column(db.Integer, db.ForeignKey('deuda.id'), nullable=False)
    monto_usd = db.Column(db.Float, nullable=False)
    monto_original = db.Column(db.Float, nullable=False) # Lo que entregó en físico
    moneda_pago = db.Column(db.String(10), nullable=False) # "USD" o "BS"
    tasa_momento = db.Column(db.Float, nullable=False)
    fecha = db.Column(db.DateTime, default=db.func.now())

class Proveedor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    telefono = db.Column(db.String(20), nullable=True)
    rubro = db.Column(db.String(100), nullable=True, default="General") 
    deuda_pendiente = db.Column(db.Float, nullable=True, default=0.0)
    cuentas_pagar = db.relationship('CuentaPorPagar', backref='proveedor', lazy=True)

class Ingreso(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(50), nullable=False)
    monto = db.Column(db.Float, nullable=False)
    metodo_pago = db.Column(db.String(50), nullable=False)
    fecha = db.Column(db.String(10), nullable=False)
    cerrado = db.Column(db.Boolean, default=False)

class Gasto(db.Model):
    """ Registro financiero de egresos o costos operativos """
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(50), nullable=False)
    monto = db.Column(db.Float, nullable=False)
    descripcion = db.Column(db.String(200))
    metodo_pago = db.Column(db.String(20), default='bs') # 'bs' o 'usd_efectivo'
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    cerrado = db.Column(db.Boolean, default=False)

class CuentaPorPagar(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    
    # Enlace real con el proveedor
    proveedor_id = db.Column(db.Integer, db.ForeignKey('proveedor.id'), nullable=False)
    
    # Control Multimoneda
    moneda = db.Column(db.String(3), nullable=False, default='USD') # 'USD' o 'BS'
    monto_original = db.Column(db.Float, nullable=False)            # El monto tal cual se negoció
    tasa_factura = db.Column(db.Float, nullable=False, default=1.0) # La tasa de cambio de ese día
    
    # Monto estandarizado en dólares para tu contabilidad global
    monto = db.Column(db.Float, nullable=False)
    
    descripcion = db.Column(db.String(200), nullable=True)
    fecha_limite = db.Column(db.String(20), nullable=True)
    pagado = db.Column(db.Boolean, default=False)

class UsuarioAdmin(db.Model):
    """ Credenciales de seguridad de acceso y preguntas de respaldo para recuperación """
    id = db.Column(db.Integer, primary_key=True)
    password_hash = db.Column(db.String(200), nullable=False)
    pregunta1 = db.Column(db.String(150), nullable=False)
    respuesta1 = db.Column(db.String(150), nullable=False)
    pregunta2 = db.Column(db.String(150), nullable=False)
    respuesta2 = db.Column(db.String(150), nullable=False)
    pregunta3 = db.Column(db.String(150), nullable=False)
    respuesta3 = db.Column(db.String(150), nullable=False)

class Configuracion(db.Model):
    """ Tabla global para almacenar variables del negocio como la tasa cambiaria """
    id = db.Column(db.Integer, primary_key=True)
    tasa_dolar = db.Column(db.Float, default=1.0)
    licencia_key = db.Column(db.String(250), nullable=True)
    fecha_vencimiento_licencia = db.Column(db.Date, nullable=True) # None = Vitalicia
    tipo_licencia = db.Column(db.String(50), default="Demo") # 'Vitalicia', 'Financiada', 'Mensual'

# ========================================================
# RUTAS DE CONTROL DE FLUJO Y SEGURIDAD (AUTENTICACIÓN)
# ========================================================

@app.route('/bienvenida', methods=['GET', 'POST'])
def bienvenida():
    conn = sqlite3.connect('negocio.db') # Cambia por el nombre exacto de tu archivo .db
    cursor = conn.cursor()
    
    # 1. Contamos cuántos usuarios existen en la tabla
    cursor.execute("SELECT COUNT(*) FROM usuarios")
    cantidad_usuarios = cursor.fetchone()[0]
    conn.close()

    if request.method == 'POST':
        usuario_ingresado = request.form.get('usuario')
        password_ingresado = request.form.get('password')

        # CASO A: No hay usuarios en el sistema todavía (Primer encendido)
        if cantidad_usuarios == 0:
            if usuario_ingresado == 'admin' and password_ingresado == 'admin':
                session['primer_inicio_autorizado'] = True
                return redirect(url_for('configurar_dueno')) # Lo mandamos a crear su cuenta
            else:
                flash("Para la configuración inicial use el usuario maestro de fábrica.", "danger")
                return render_template('bienvenida.html')

        # CASO B: Ya el sistema está configurado con usuarios reales
        else:
            if usuario_ingresado == 'admin' and password_ingresado == 'admin':
                flash("La cuenta de fábrica 'admin' ha sido deshabilitada por seguridad.", "danger")
                return render_template('bienvenida.html')
            
            # Buscamos al usuario real en la base de datos
            conn = sqlite3.connect('negocio.db')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM usuarios WHERE usuario = ? AND password = ?", (usuario_ingresado, password_ingresado))
            user = cursor.fetchone()
            conn.close()

            if user:
                session['logueado'] = True
                session['usuario_actual'] = usuario_ingresado
                return redirect(url_for('menu_seleccion')) # ¡Entra al sistema!
            else:
                flash("Usuario o contraseña incorrectos.", "danger")

    return render_template('bienvenida.html')

@app.route('/configurar_dueno', methods=['GET', 'POST'])
def configurar_dueno():
    # Evita que cualquiera entre a esta ruta si no puso admin/admin primero
    if not session.get('primer_inicio_autorizado'):
        return redirect(url_for('bienvenida'))

    if request.method == 'POST':
        nuevo_usuario = request.form.get('nuevo_usuario')
        nuevo_password = request.form.get('nuevo_password')

        if nuevo_usuario == 'admin':
            flash("No puede usar 'admin' como usuario real. Elija otro nombre.", "danger")
            return render_template('configurar_dueno.html')

        # Usamos un bloque try/except para atrapar el error de duplicado sin que se caiga el sistema
        try:
            conn = sqlite3.connect('negocio.db')
            cursor = conn.cursor()
            cursor.execute("INSERT INTO usuarios (usuario, password) VALUES (?, ?)", (nuevo_usuario, nuevo_password))
            conn.commit()
            conn.close()

            # Limpiamos la sesión temporal y lo logueamos oficialmente
            session.pop('primer_inicio_autorizado', None)
            session['logueado'] = True
            session['usuario_actual'] = nuevo_usuario

            return redirect(url_for('menu_seleccion')) # Va directo al menú principal ya protegido

        except sqlite3.IntegrityError:
            #  Si el usuario ya existe en la base de datos, entra aquí en lugar de dar error
            flash(f"El usuario '{nuevo_usuario}' ya está registrado en el sistema. Intente con otro nombre.", "danger")
            return render_template('configurar_dueno.html')

    return render_template('configurar_dueno.html')
    
@app.route('/menu')
def menu_seleccion():
    """ Pantalla intermedia para elegir entre ventas o finanzas después del login """
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
    return render_template('menu.html')

@app.route('/logout')
def logout():
    """ Destruye la sesión del usuario de forma segura """
    session.pop('logueado', None)
    flash("Has cerrado sesión correctamente.", "info")
    return redirect(url_for('bienvenida'))

@app.route('/registrar_admin', methods=['POST'])
def registrar_admin():
    """ Genera el único usuario maestro de la base de datos con sus preguntas secretas """
    if UsuarioAdmin.query.first():
        flash("El administrador ya está registrado.", "danger")
        return redirect(url_for('bienvenida'))
    
    password = request.form.get('password')
    p1 = request.form.get('pregunta1')
    r1 = request.form.get('respuesta1', '').strip().lower()
    p2 = request.form.get('pregunta2')
    r2 = request.form.get('respuesta2', '').strip().lower()
    p3 = request.form.get('pregunta3')
    r3 = request.form.get('respuesta3', '').strip().lower()

    if not password or not r1 or not r2 or not r3:
        flash("Todos los campos obligatorios deben completarse.", "warning")
        return redirect(url_for('bienvenida'))

    hashed_pw = generate_password_hash(password)
    
    nuevo_usuario = UsuarioAdmin(
        password_hash=hashed_pw,
        pregunta1=p1, respuesta1=r1,
        pregunta2=p2, respuesta2=r2,
        pregunta3=p3, respuesta3=r3
    )
    db.session.add(nuevo_usuario)
    db.session.commit()
    
    session['logueado'] = True
    flash("¡Administrador registrado con éxito! Bienvenido.", "success")
    return redirect(url_for('index'))

@app.route('/recuperar_password', methods=['POST'])
def recuperar_password():
    """ Evalúa las respuestas de seguridad para autorizar el reseteo de la clave corporativa """
    admin = UsuarioAdmin.query.first()
    if not admin:
        return redirect(url_for('bienvenida'))
        
    r1 = request.form.get('respuesta1', '').strip().lower()
    r2 = request.form.get('respuesta2', '').strip().lower()
    r3 = request.form.get('respuesta3', '').strip().lower()
    nueva_pw = request.form.get('nueva_password')
    
    if r1 == admin.respuesta1 and r2 == admin.respuesta2 and r3 == admin.respuesta3:
        admin.password_hash = generate_password_hash(nueva_pw)
        db.session.commit()
        flash("Contraseña restablecida con éxito. Inicia sesión.", "success")
    else:
        flash("Respuestas incorrectas. Verificación de identidad denegada.", "danger")
        
    return redirect(url_for('bienvenida'))

# ========================================================
# PANEL PRINCIPAL (INDEX) Y PROCESOS OPERATIVOS DEL NEGOCIO
# ========================================================

@app.route('/')
@licencia_requerida
def index():
    if not session.get('logueado'):
        return redirect(url_for('bienvenida'))
        
    config = Configuracion.query.first()
    if not config:
        config = Configuracion(tasa_dolar=1.0)
        db.session.add(config)
        db.session.commit()

    tasa_actual = float(config.tasa_dolar)

    ingresos_lista = Ingreso.query.filter_by(cerrado=False).order_by(Ingreso.id.desc()).all()
    gastos_lista = Gasto.query.filter_by(cerrado=False).order_by(Gasto.id.desc()).all()
    
    # -------------------------------------------------------------
    # CONTABILIZACIÓN DE INGRESOS (En Caja)
    # -------------------------------------------------------------
    ingresos_en_dolares = 0.0
    ingresos_en_bs = 0.0

    for i in ingresos_lista:
        monto_registro = float(i.monto or 0.0)
        metodo = str(i.metodo_pago or "").lower().strip()
        
        es_divisa = ("dolar" in metodo or "dólar" in metodo or "divisa" in metodo or "zelle" in metodo or "usd" in metodo or "💵" in metodo)
        if es_divisa:
            ingresos_en_dolares += monto_registro
        else:
            # Si el ingreso se registró en Bs originalmente, lo sumamos directo a Bs
            ingresos_en_bs += monto_registro

    # -------------------------------------------------------------
    # CONTABILIZACIÓN DE GASTOS (En Caja)
    # -------------------------------------------------------------
    gastos_en_dolares = 0.0
    gastos_en_bs = 0.0

    for g in gastos_lista:
        monto_gasto = float(g.monto or 0.0)
        metodo_g = str(g.metodo_pago or "").lower().strip()
        
        es_divisa_g = ("dolar" in metodo_g or "dólar" in metodo_g or "divisa" in metodo_g or "zelle" in metodo_g or "usd" in metodo_g or "💵" in metodo_g)
        if es_divisa_g:
            gastos_en_dolares += monto_gasto
        else:
            # Si el gasto se registró en Bs originalmente, lo sumamos directo a Bs
            gastos_en_bs += monto_gasto

    # -------------------------------------------------------------
    #  GANANCIA NETA (Resta directa por moneda independiente)
    # -------------------------------------------------------------
    ganancia_neta_usd = ingresos_en_dolares - gastos_en_dolares
    ganancia_neta_bs = ingresos_en_bs - gastos_en_bs
    # -------------------------------------------------------------

    productos = Producto.query.all()
    clientes = Cliente.query.all()
    proveedores = Proveedor.query.all()
    cuentas_pagar = CuentaPorPagar.query.all()

    dinero_debe_usd = sum(float(c.monto or 0.0) for c in cuentas_pagar)
    # Calculamos el equivalente total en bolívares de esa deuda
    dinero_debe_bs = dinero_debe_usd * tasa_actual

    productos_ordenados = sorted(productos, key=lambda p: p.ventas_totales or 0, reverse=True)
    productos_estrella = [p for p in productos_ordenados if (p.ventas_totales or 0) > 0][:3]
    productos_frios = [p for p in productos if (p.ventas_totales or 0) == 0]

    return render_template('index.html',
                           productos=productos,
                           clientes=clientes,
                           proveedores=proveedores,
                           ingresos=ingresos_lista,
                           gastos=gastos_lista,
                           cuentas_pagar=cuentas_pagar,
                           dinero_generado=ingresos_en_dolares,
                           
                           # VARIABLES CLAVE PARA TU CUADRO ROJO:
                           dinero_debe=dinero_debe_usd,  # Envía el monto limpio en $
                           dinero_debe_bs=dinero_debe_bs, # Envía el monto convertido a Bs
                           
                           productos_estrella=productos_estrella,
                           productos_frios=productos_frios,
                           tasa_dolar=tasa_actual,
                           ingresos_en_dolares=ingresos_en_dolares,
                           ingresos_en_bs=ingresos_en_bs,
                           gastos_en_dolares=gastos_en_dolares,
                           gastos_en_bs=gastos_en_bs,
                           ganancia_neta_usd=ganancia_neta_usd,
                           ganancia_neta_bs=ganancia_neta_bs)

@app.route('/actualizar_tasa', methods=['POST'])
def actualizar_tasa():
    """ Modifica el multiplicador de conversión Bs/USD de forma global """
    if not session.get('logueado'): return redirect(url_for('bienvenida'))
    nueva_tasa = float(request.form.get('tasa_dolar', 1.0))
    config = Configuracion.query.first()
    if config:
        config.tasa_dolar = nueva_tasa
        db.session.commit()
        flash("Tasa del dólar actualizada correctamente.", "success")
    return redirect(url_for('index'))


@app.route('/registrar_ingreso', methods=['POST'])
def registrar_ingreso():
    """ Procesa tanto ventas directas de inventario como inyecciones de dinero externas """
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
        
    tipo_formulario = request.form.get('tipo_ingreso')
    metodo = request.form.get('metodo_pago')
    
    config = Configuracion.query.first()
    tasa = config.tasa_dolar if config else 1.0

    import datetime
    fecha_hoy = datetime.date.today().strftime("%Y-%m-%d")

    if tipo_formulario == "producto":
        # 🛒 PROCESAR VENTA DE PRODUCTOS DEL INVENTARIO (Queda igual, ya funciona bien)
        productos_ids = request.form.getlist('productos[]')
        cantidades = request.form.getlist('cantidades[]')
        
        total_venta_usd = 0.0
        detalles_venta = []

        for p_id, cant_str in zip(productos_ids, cantidades):
            if not p_id or not cant_str:
                continue
            
            cant = int(cant_str)
            if cant <= 0:
                continue

            prod = Producto.query.get(int(p_id))
            if prod:
                if prod.stock < cant:
                    flash(f"Error: Stock insuficiente para el producto '{prod.nombre}'.", "danger")
                    return redirect(url_for('index'))
                
                prod.stock -= cant
                prod.ventas_totales = (prod.ventas_totales or 0) + cant
                total_venta_usd += (prod.precio_venta * cant)
                detalles_venta.append(f"{prod.nombre} (x{cant})")

        if total_venta_usd > 0:
            descripcion_final = f"Venta: {', '.join(detalles_venta)}"
            nuevo_ingreso = Ingreso(
                tipo=descripcion_final,
                monto=total_venta_usd,
                metodo_pago=metodo,
                fecha=fecha_hoy,
                cerrado=False
            )
            db.session.add(nuevo_ingreso)
            db.session.commit() # Guardamos en la base de datos
            flash("Venta registrada con éxito.", "success")
        else:
            flash("No seleccionaste ningún producto válido.", "warning")
            return redirect(url_for('index'))

    else:
        # PROCESAR INGRESO EXTERNO 
        monto_ext_str = request.form.get('monto_externo')
        concepto = request.form.get('concepto_externo') or "Otros Ingresos"
        
        if monto_ext_str:
            monto_ext = float(monto_ext_str)
            
            # AQUÍ ESTÁ EL TRUCO INTELIGENTE:
            if metodo == "bs":
                # Si el usuario seleccionó Bolívares, calculamos su equivalente en USD dividiendo por la tasa
                # Así la columna 'monto' guarda el valor real equivalente en dólares
                monto_final_usd = monto_ext / tasa if tasa > 0 else monto_ext
                # Le añadimos una nota al concepto para que en tus tablas recuerdes cuántos Bs eran originalmente
                concepto_final = f"{concepto} ({monto_ext:.2f} Bs)"
            else:
                # Si seleccionó dólares en efectivo, el monto base es exactamente el valor plano digitado
                monto_final_usd = monto_ext
                concepto_final = concepto

            nuevo_ingreso = Ingreso(
                tipo=concepto_final,
                monto=monto_final_usd, # Guardamos el contravalor correcto en USD base
                metodo_pago=metodo,     # Guarda 'bs' o 'usd_efectivo' para saber a qué cuenta va
                fecha=fecha_hoy,
                cerrado=False
            )
            db.session.add(nuevo_ingreso)
            db.session.commit() # Guardamos en la base de datos
            flash("Ingreso externo registrado con éxito.", "success")
        else:
            flash("Monto externo no válido.", "warning")
            return redirect(url_for('index'))

    return redirect(url_for('index'))

@app.route('/registrar_gasto', methods=['POST'])
def registrar_gasto():
    """ Procesa el registro de un egreso o gasto del negocio """
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
        
    # Intentamos leer 'concepto', si no existe 'descripcion', y si no 'tipo'.
    concepto = request.form.get('concepto') or request.form.get('descripcion') or request.form.get('tipo') or "Gasto General"
    
    metodo = request.form.get('metodo_pago')
    monto_str = request.form.get('monto')

    # CAPTURAMOS LA TASA DINÁMICA
    tasa_raw = request.form.get('tasa_gasto')
    tasa_gasto = float(tasa_raw) if tasa_raw and tasa_raw.strip() != "" else 1.0

    import datetime
    fecha_hoy = datetime.date.today() 

    if monto_str:
        monto_original = float(monto_str)
        
        # VALIDACIÓN MULTIMONEDA: Si es en bolívares, hacemos la conversión a USD
        if metodo == 'bs':
            monto_gasto = monto_original / tasa_gasto
        else:
            monto_gasto = monto_original # Si es dólares, pasa directo

        # Creamos el registro en la base de datos con el monto ya en USD
        nuevo_gasto = Gasto(
            tipo=concepto,          # Se guarda en el campo 'tipo' de tu tabla
            monto=round(monto_gasto, 2), # Redondeamos a 2 decimales para evitar decimales infinitos     
            metodo_pago=metodo,     
            fecha=fecha_hoy,        
            cerrado=False
        )
        db.session.add(nuevo_gasto)
        db.session.commit()
        flash("Gasto registrado con éxito.", "success")
    else:
        flash("Monto de gasto no válido.", "warning")

    return redirect(url_for('index'))

@app.route('/agregar_cuenta_pagar', methods=['POST'])
def agregar_cuenta_pagar():
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
        
    proveedor_id = int(request.form.get('proveedor_id'))
    moneda = request.form.get('moneda')
    monto_original = float(request.form.get('monto_original', 0.0))
    descripcion = request.form.get('descripcion')
    fecha_limite = request.form.get('fecha_limite')

    # 1. Traemos la tasa de configuración de tu negocio (ej: 800 Bs)
    config = Configuracion.query.first()
    tasa_negocio = float(config.tasa_dolar) if config else 1.0

    tasa_raw = request.form.get('tasa_factura')
    tasa_factura = float(tasa_raw) if tasa_raw and tasa_raw.strip() != "" else 1.0

    # 2. Convertimos el monto de la deuda usando la TASA DE TU NEGOCIO si es en BS
    # Si debes 80.000 Bs, se divide entre tu tasa de 800 Bs -> Deuda real de 100 USD en tu sistema
    if moneda == 'BS':
        monto_usd = monto_original / tasa_negocio
    else:
        monto_usd = monto_original
        tasa_factura = 1.0  # Si es dólares, la tasa interna es 1

    # Guardamos la descripción sumándole un detalle para que sepas a qué tasa se calculó originalmente
    if moneda == 'BS':
        descripcion_extra = f" [Facturado a tasa {tasa_factura} Bs]"
        descripcion = (descripcion or "") + descripcion_extra

    nueva_deuda = CuentaPorPagar(
        proveedor_id=proveedor_id,
        moneda=moneda,
        monto_original=round(monto_original, 2),
        tasa_factura=round(tasa_factura, 2), # Guardamos la tasa de la factura por si necesitas consultarla
        monto=round(monto_usd, 2),            # ¡Monto en USD calculado a la tasa de tu caja!
        descripcion=descripcion,
        fecha_limite=fecha_limite,
        pagado=False
    )
    
    db.session.add(nueva_deuda)
    db.session.commit()
    
    flash("Cuenta por pagar registrada de forma inteligente.", "success")
    return redirect(url_for('finanzas'))

@app.route('/abonar_cuenta_pagar/<int:id>', methods=['POST'])
def abonar_cuenta_pagar(id):
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
    
    cuenta = CuentaPorPagar.query.get_or_404(id)
    
    moneda_pago = request.form.get('moneda_pago')
    monto_abonado = float(request.form.get('monto_abonado', 0.0))
    tasa_pago = float(request.form.get('tasa_pago', 1.0))
    
    # 1. Obtenemos la tasa de configuración de tu negocio (ej: 800 Bs)
    config = Configuracion.query.first()
    tasa_negocio = float(config.tasa_dolar) if config else 1.0
    
    # A) Conversión para el PROVEEDOR (Tasa acordada, ej: 730) -> Da 109.59 USD
    if moneda_pago == 'BS':
        abono_usd_proveedor = monto_abonado / tasa_pago
    else:
        abono_usd_proveedor = monto_abonado

    # B) Conversión para TU CAJA (Tasa del negocio, ej: 800) -> Da 100.00 USD
    if moneda_pago == 'BS':
        abono_usd_caja = monto_abonado / tasa_negocio
    else:
        abono_usd_caja = monto_abonado
        tasa_pago = 1.0

    # Validamos contra el saldo en el sistema usando el impacto real al proveedor
    if abono_usd_proveedor > (cuenta.monto + 0.01):
        flash(f"Error: El abono (${abono_usd_proveedor:.2f}) supera el saldo pendiente (${cuenta.monto:.2f}).", "danger")
        return redirect(url_for('finanzas'))

    # Restamos del compromiso pendiente el equivalente real acordado con el proveedor
    cuenta.monto = round(cuenta.monto - abono_usd_proveedor, 2)
    
    # Si ya se cubrió la deuda por completo, la marcamos como resuelta
    if cuenta.monto <= 0.05:
        cuenta.monto = 0.0
        cuenta.pagado = True
        flash(f"¡Felicidades! Deuda liquidada con {cuenta.proveedor.nombre if cuenta.proveedor else 'Proveedor'}.", "success")
    else:
        flash(f"Abono registrado por {monto_abonado} {moneda_pago}.", "success")

    # 2. REGISTRAMOS EL GASTO HOMOGENEIZADO A LA TASA DE TU CAJA
    detalles_gasto = f"Pago/Abono de deuda a {cuenta.proveedor.nombre if cuenta.proveedor else 'Proveedor'}: {monto_abonado:.2f} {moneda_pago}"
    if moneda_pago == 'BS':
        detalles_gasto += f" (Pactado a tasa: {tasa_pago} | Registrado a tasa caja: {tasa_negocio})"
        
    nuevo_gasto = Gasto(
        tipo="Proveedor (Pago de Deuda)",
        monto=round(abono_usd_caja, 2),  # <--- ¡CORREGIDO! Aquí caen tus 100 USD exactos
        descripcion=detalles_gasto,
        metodo_pago="Efectivo Dólar" if moneda_pago == 'USD' else "Transferencia/Pago Móvil Bs",
        cerrado=False
    )
    db.session.add(nuevo_gasto)
    db.session.commit()
    
    return redirect(url_for('finanzas'))

@app.route('/abonar_cliente', methods=['POST'])
def abonar_cliente():
    """ Permite amortizar o liquidar la deuda de un cliente, actualizando su nivel """
    if not session.get('logueado'): return redirect(url_for('bienvenida'))
    clie_id = int(request.form.get('cliente_id'))
    monto_abono = float(request.form.get('monto', 0))
    metodo = request.form.get('metodo_pago')

    clie = Cliente.query.get(clie_id)
    if clie and monto_abono > 0:
        clie.deuda = max(0.0, clie.deuda - monto_abono)
        if clie.deuda == 0:
            clie.nivel = 'Al Día'
            
        nuevo_ingreso = Ingreso(monto=monto_abono, descripcion=f"Abono de cliente: {clie.nombre}", metodo_pago=metodo)
        db.session.add(nuevo_ingreso)
        db.session.commit()
        flash("Abono procesado correctamente.", "success")
    return redirect(url_for('index'))

@app.route('/pagar_cuenta', methods=['POST'])
def pagar_cuenta():
    """ Liquida total o parcialmente una obligación con tasa personalizada e inyecta el gasto en la caja activa """
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
    
    # SALVAVIDAS: Si la base de datos es nueva, aseguramos que exista la configuración de la tasa
    config = Configuracion.query.first()
    if not config:
        config = Configuracion(tasa_dolar=1.0)
        db.session.add(config)
        db.session.commit()
    
    cuenta_id = int(request.form.get('cuenta_id'))
    monto_a_pagar_usd = float(request.form.get('monto_a_pagar', 0.0))
    moneda_pago = request.form.get('moneda_pago')  # 'USD' o 'BS'
    tasa_pago = float(request.form.get('tasa_pago', 1.0)) 

    cuenta = CuentaPorPagar.query.get(cuenta_id)
    if cuenta:
        if monto_a_pagar_usd <= 0 or monto_a_pagar_usd > cuenta.monto:
            flash("Monto de abono inválido.", "danger")
            return redirect(url_for('index'))

        prov = Proveedor.query.get(cuenta.proveedor_id)
        
        # 1. Calcular el monto real que saldrá de los Gastos según la moneda
        if moneda_pago == 'BS':
            monto_gasto_real = monto_a_pagar_usd * tasa_pago
            metodo_gasto = f"Transferencia/Pago Móvil Bs (Tasa: {tasa_pago})"
            detalle_gasto = f"Abono de ${monto_a_pagar_usd:.2f} a deuda de {prov.nombre if prov else 'Proveedor'} pagado en Bs a tasa {tasa_pago}. (Ref: {cuenta.descripcion})"
        else:
            monto_gasto_real = monto_a_pagar_usd
            metodo_gasto = "Efectivo Dólar / Divisa"
            detalle_gasto = f"Abono de ${monto_a_pagar_usd:.2f} a deuda de {prov.nombre if prov else 'Proveedor'} pagado en USD. (Ref: {cuenta.descripcion})"

        # 2. Registrar el Gasto automático (CORREGIDO: Añadido el campo 'tipo')
        nuevo_gasto = Gasto(
            tipo="Proveedor",  # Puedes cambiarlo por "Variable" o "Otros" si usas otra categoría en tus modelos
            monto=monto_gasto_real, 
            descripcion=detalle_gasto, 
            metodo_pago=metodo_gasto,
            cerrado=False
        )
        db.session.add(nuevo_gasto)
        
        # 3. Restar el abono a la deuda general del proveedor
        if prov:
            prov.deuda_pendiente = max(0.0, prov.deuda_pendiente - monto_a_pagar_usd)
            
        # 4. Restar el abono a esta factura/cuenta específica
        cuenta.monto -= monto_a_pagar_usd

        if cuenta.monto <= 0.01:
            db.session.delete(cuenta)
            flash("Obligación liquidada por completo.", "success")
        else:
            flash(f"Abono de ${monto_a_pagar_usd:.2f} procesado con éxito. Saldo restante: ${cuenta.monto:.2f}", "success")
            
        db.session.commit()
        
    return redirect(url_for('index'))

# ========================================================
# GENERADOR DE CIERRE DE CAJA AUTOMÁTICO A EXCEL
# ========================================================

@app.route('/cierre_caja')
def closure_caja():
    """ Genera el libro Excel en una sola hoja con el formato de 4 columnas operativas paralelas """
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))

    ingresos_dia = Ingreso.query.filter_by(cerrado=False).all()
    gastos_dia = Gasto.query.filter_by(cerrado=False).all()

    if not ingresos_dia and not gastos_dia:
        flash("No hay movimientos activos hoy para realizar un cierre.", "warning")
        return redirect(url_for('index'))

    # Obtener la tasa del negocio
    config = Configuracion.query.first()
    tasa_del_dia = float(config.tasa_dolar) if config else 1.0

    # 1. CLASIFICACIÓN DE MOVIMIENTOS EN PYTHON
    ingresos_bs = []
    ingresos_usd = []
    for ing in ingresos_dia:
        metodo = (ing.metodo_pago or 'BS').upper()
        if 'USD' in metodo or 'DOLAR' in metodo or 'EFECTIVO DÓLAR' in metodo:
            ingresos_usd.append(ing)
        else:
            ingresos_bs.append(ing)

    gastos_dia_lista = []
    gastos_usd = []
    # Cambié el nombre de la variable temporal a 'gas' para evitar confusiones
    for gas in gastos_dia:
        metodo = (gas.metodo_pago or 'BS').upper()
        if 'USD' in metodo or 'DOLAR' in metodo or 'EFECTIVO DÓLAR' in metodo:
            gastos_usd.append(gas)
        else:
            gastos_dia_lista.append(gas)

    max_filas = max(len(ingresos_bs), len(ingresos_usd), len(gastos_dia_lista), len(gastos_usd), 1)

    # 2. CREACIÓN DEL ARCHIVO EXCEL
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cierre de Caja"
    ws.views.sheetView[0].showGridLines = True

    # Estilos corporativos
    font_cabecera = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_subcabecera = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_datos = Font(name="Segoe UI", size=10)
    font_totales_lbl = Font(name="Segoe UI", size=10, bold=True, color="1F497D")
    font_totales_val = Font(name="Segoe UI", size=11, bold=True)
    
    fill_ingresos = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_gastos = PatternFill(start_color="A62B2B", end_color="A62B2B", fill_type="solid")
    fill_totales = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    # 3. ESCRIBIR CABECERAS PRINCIPALES (Fila 7)
    headers = [
        ("INGRESOS TOTALES BS", fill_ingresos), ("RAZON", fill_ingresos),
        ("INGRESOS TOTALES $", fill_ingresos), ("RAZON", fill_ingresos),
        ("GASTOS TOTALES BS", fill_gastos), ("RAZON", fill_gastos),
        ("GASTOS TOTALES $", fill_gastos), ("RAZON", fill_gastos),
        ("TASA DEL DIA", fill_totales), ("GANANCIA TOTAL BS", fill_totales), ("GANANCIA TOTAL $", fill_totales)
    ]

    for col_idx, (texto, fill) in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=texto)
        cell.font = font_cabecera
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subheaders = {
        1: "MONTO EN BS", 2: "QUE SE VENDIO",
        3: "MONTO EN DOLARES", 4: "QUE SE VENDIO",
        5: "MONTO EN BS", 6: "PORQUE ESE GASTO",
        7: "MONTO EN DOLARES", 8: "PORQUE ESE GASTO"
    }
    for col_idx, texto in subheaders.items():
        cell = ws.cell(row=8, column=col_idx, value=texto)
        cell.font = font_subcabecera
        cell.fill = headers[col_idx-1][1]
        cell.alignment = Alignment(horizontal="center")

    # 4. VOLCADO DE LAS CUATRO COLUMNAS EN PARALELO
    start_row = 9
    for i in range(max_filas):
        current_row = start_row + i
        row_fill = fill_zebra if i % 2 == 1 else None

        # --- Ingresos Bs (Cols 1 y 2) ---
        if i < len(ingresos_bs):
            ing = ingresos_bs[i]
            c_monto = ws.cell(row=current_row, column=1, value=ing.monto)
            c_monto.number_format = '#,##0.00'
            
            # CONTROL DE SEGURIDAD PARA LA DESCRIPCIÓN DEL INGRESO
            razon_ing = getattr(ing, 'descripcion', None) or getattr(ing, 'tipo', None) or "Venta Mercancía"
            ws.cell(row=current_row, column=2, value=razon_ing)
        
        # --- Ingresos USD (Cols 3 y 4) ---
        if i < len(ingresos_usd):
            ing = ingresos_usd[i]
            c_monto = ws.cell(row=current_row, column=3, value=ing.monto)
            c_monto.number_format = '$#,##0.00'
            
            # CONTROL DE SEGURIDAD PARA LA DESCRIPCIÓN DEL INGRESO
            razon_ing = getattr(ing, 'descripcion', None) or getattr(ing, 'tipo', None) or "Venta Mercancía"
            ws.cell(row=current_row, column=4, value=razon_ing)

        # --- Gastos Bs (Cols 5 y 6) ---
        if i < len(gastos_dia_lista):
            gas = gastos_dia_lista[i]
            c_monto = ws.cell(row=current_row, column=5, value=gas.monto)
            c_monto.number_format = '#,##0.00'
            
            razon_gas = getattr(gas, 'descripcion', None) or getattr(gas, 'tipo', None) or "Gasto General"
            ws.cell(row=current_row, column=6, value=razon_gas)

        # --- Gastos USD (Cols 7 y 8) ---
        if i < len(gastos_usd):
            gas = gastos_usd[i]
            c_monto = ws.cell(row=current_row, column=7, value=gas.monto)
            c_monto.number_format = '$#,##0.00'
            
            razon_gas = getattr(gas, 'descripcion', None) or getattr(gas, 'tipo', None) or "Gasto General"
            ws.cell(row=current_row, column=8, value=razon_gas)

        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_datos
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    end_row_data = start_row + max_filas - 1

    # 5. PANEL LATERAL ESTÁTICO (Fila 11 Fija)
    panel_row = 11
    
    c_tasa = ws.cell(row=panel_row, column=9, value=tasa_del_dia)
    c_tasa.font = font_totales_val
    c_tasa.number_format = '0.00'
    c_tasa.alignment = Alignment(horizontal="center")
    c_tasa.border = thin_border
    
    formula_ganancia_bs = f"=(SUM(A{start_row}:A{end_row_data})-SUM(E{start_row}:E{end_row_data}))+(K{panel_row}*I{panel_row})"
    c_gan_bs = ws.cell(row=panel_row, column=10, value=formula_ganancia_bs)
    c_gan_bs.font = font_totales_val
    c_gan_bs.number_format = '#,##0.00" Bs"'
    c_gan_bs.border = thin_border
    
    formula_ganancia_usd = f"=SUM(C{start_row}:C{end_row_data})-SUM(G{start_row}:G{end_row_data})"
    c_gan_usd = ws.cell(row=panel_row, column=11, value=formula_ganancia_usd)
    c_gan_usd.font = font_totales_val
    c_gan_usd.number_format = '$#,##0.00'
    c_gan_usd.border = thin_border

    # Fila de Cierre Totalizadores por debajo de las tablas operativas
    tot_row = end_row_data + 1
    ws.cell(row=tot_row, column=1, value=f"=SUM(A{start_row}:A{end_row_data})").number_format = '#,##0.00'
    ws.cell(row=tot_row, column=3, value=f"=SUM(C{start_row}:C{end_row_data})").number_format = '$#,##0.00'
    ws.cell(row=tot_row, column=5, value=f"=SUM(E{start_row}:E{end_row_data})").number_format = '#,##0.00'
    ws.cell(row=tot_row, column=7, value=f"=SUM(G{start_row}:G{end_row_data})").number_format = '$#,##0.00'
    
    for c_idx in [1, 3, 5, 7]:
        ws.cell(row=tot_row, column=c_idx).font = font_totales_lbl
        ws.cell(row=tot_row, column=c_idx).border = thin_border

    # Autoajuste de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 16)

    # 6. CIERRE DE MOVIMIENTOS HISTÓRICOS
    for ing in ingresos_dia: 
        ing.cerrado = True
    for gas in gastos_dia: 
        gas.cerrado = True
    db.session.commit()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    fecha_str = datetime.now().strftime("%d-%m-%Y")
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Cierre_Caja_{fecha_str}.xlsx"
    )

# Modifica tu ruta original /finanzas para que busque a los clientes y productos
@app.route('/finanzas')
@licencia_requerida
def finanzas():
    if not session.get('logueado'): return redirect(url_for('bienvenida'))
    
    config = Configuracion.query.first()
    tasa_actual = float(config.tasa_dolar) if config else 1.0

    proveedores = Proveedor.query.all()
    cuentas_pagar = CuentaPorPagar.query.all()
    dinero_debe_usd = sum(float(c.monto or 0.0) for c in cuentas_pagar)

    # NUEVAS CONSULTAS COMPLEMENTARIAS
    clientes = Cliente.query.all()
    todos_productos = Producto.query.all()

    return render_template('finanzas.html',
                           proveedores=proveedores,
                           cuentas_pagar=cuentas_pagar,
                           dinero_debe=dinero_debe_usd,
                           tasa_dolar=tasa_actual,
                           fecha_actual=datetime.now().date(),
                           clientes=clientes,
                           todos_productos=todos_productos)


@app.route('/crear_deuda_externa', methods=['POST'])
def crear_deuda_externa():
    nombre = request.form.get('nombre_cliente')
    telefono = request.form.get('telefono_cliente')
    monto = float(request.form.get('monto_deuda', 0.0))
    moneda = request.form.get('moneda')
    descripcion = request.form.get('descripcion')

    config = Configuracion.query.first()
    tasa = float(config.tasa_dolar) if config else 1.0
    
    # Dolarizar si viene en Bs usando la tasa general del mostrador
    monto_usd = monto / tasa if moneda == 'BS' else monto

    # Verificar o crear cliente unico
    cliente = Cliente.query.filter_by(nombre=nombre).first()
    if not cliente:
        cliente = Cliente(nombre=nombre, telefono=telefono, deuda_total=0.0)
        db.session.add(cliente)
        db.session.commit()

    nueva_deuda = Deuda(
        cliente_id=cliente.id,
        tipo="Externa",
        monto_inicial=monto_usd,
        saldo_pendiente=monto_usd,
        descripcion=descripcion
    )
    cliente.deuda_total += monto_usd
    db.session.add(nueva_deuda)
    db.session.commit()
    
    flash("Deuda externa registrada y congelada en dólares.", "success")
    return redirect(url_for('finanzas'))

@app.route('/crear_deuda_productos', methods=['GET', 'POST'])
@licencia_requerida
def crear_deuda_productos():
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
    
    nombre_cliente = request.form.get('nombre_cliente')
    telefono = request.form.get('telefono_cliente')
    
    productos_ids = request.form.getlist('productos[]')
    cantidades = request.form.getlist('cantidades[]')
    
    monto_acumulado_usd = 0.0
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    detalles_texto = f"Compra inicial ({fecha_actual}):\n"
    
    productos_procesados = 0

    for p_id, cant in zip(productos_ids, cantidades):
        if not p_id or not cant: 
            continue
        
        producto = Producto.query.get(int(p_id))
        cantidad = int(cant)
        
        if producto and producto.stock >= cantidad:
            producto.stock -= cantidad
            subtotal = producto.precio_venta * cantidad
            monto_acumulado_usd += subtotal
            detalles_texto += f"- {cantidad} un. de {producto.nombre} (${producto.precio_venta:.2f} c/u)\n"
            productos_procesados += 1

    if productos_procesados == 0:
        detalles_texto = f"Registro de deuda creado el {fecha_actual} sin especificación de artículos."

    # 1. BUSCAR O CREAR EL CLIENTE (Sin pasar 'descripcion' al constructor)
    cliente = Cliente.query.filter_by(nombre=nombre_cliente).first()
    
    if not cliente:
        # Creamos al cliente de forma limpia solo con los campos estándar
        cliente = Cliente(
            nombre=nombre_cliente,
            telefono=telefono,
            deuda_total=round(monto_acumulado_usd, 2)
        )
        db.session.add(cliente)
    else:
        cliente.deuda_total = (cliente.deuda_total or 0.0) + round(monto_acumulado_usd, 2)

    # 2. ASIGNACIÓN SEGURA DEL TEXTO (Detecta qué campo existe en tu modelo Cliente)
    if hasattr(cliente, 'descripcion'):
        cliente.descripcion = (cliente.descripcion + f"\n\n{detalles_texto}") if cliente.descripcion else detalles_texto
    elif hasattr(cliente, 'notas'):
        cliente.notas = (cliente.notas + f"\n\n{detalles_texto}") if cliente.notas else detalles_texto
    elif hasattr(cliente, 'historial'):
        cliente.historial = (cliente.historial + f"\n\n{detalles_texto}") if cliente.historial else detalles_texto

    db.session.commit() # Guardamos para generar/obtener cliente.id

    # 3. CREAR REGISTRO EN LA TABLA DEUDA (Donde la descripción SÍ es válida)
    nueva_deuda = Deuda(
        cliente_id=cliente.id,
        tipo="Productos",
        monto_inicial=round(monto_acumulado_usd, 2),
        saldo_pendiente=round(monto_acumulado_usd, 2),
        descripcion=detalles_texto,
        estado="Pendiente"
    )
    db.session.add(nueva_deuda)
    db.session.commit()
    
    flash("Cuenta por cobrar consolidada con éxito.", "success")
    return redirect(url_for('finanzas'))

@app.route('/anexar_productos_deuda/<int:cliente_id>', methods=['POST'])
@licencia_requerida
def anexar_productos_deuda(cliente_id):
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
    
    cliente = Cliente.query.get(cliente_id)
    producto_id = request.form.get('producto_id')
    cantidad = int(request.form.get('cantidad', 1))
    
    producto = Producto.query.get(producto_id)
    
    if cliente and producto:
        if producto.stock >= cantidad:
            # 1. Rebozar / Descontar stock
            producto.stock -= cantidad
            costo_adicional = producto.precio_venta * cantidad
            
            # 2. Incrementar deuda total del cliente
            cliente.deuda_total = (cliente.deuda_total or 0.0) + costo_adicional
            
            # 3. Preparar la nueva línea de texto para el historial
            fecha_str = datetime.now().strftime('%d/%m/%Y %H:%M')
            linea_anexo = f"- Anexo ({fecha_str}): {cantidad} un. de {producto.nombre} (+${costo_adicional:.2f})\n"
            
            # A) ACTUALIZAR EN EL OBJETO CLIENTE
            if hasattr(cliente, 'descripcion') and cliente.descripcion:
                cliente.descripcion += f"\n{linea_anexo}"
            elif hasattr(cliente, 'notas') and cliente.notas:
                cliente.notas += f"\n{linea_anexo}"
            else:
                cliente.descripcion = linea_anexo

            # B) ACTUALIZAR EN LA TABLA DEUDA
            deuda_activa = Deuda.query.filter_by(cliente_id=cliente_id, estado="Pendiente").first()
            if deuda_activa:
                deuda_activa.saldo_pendiente = (deuda_activa.saldo_pendiente or 0.0) + costo_adicional
                if deuda_activa.descripcion:
                    deuda_activa.descripcion += f"\n{linea_anexo}"
                else:
                    deuda_activa.descripcion = linea_anexo
            else:
                # Si no existía un registro en Deuda, se crea
                nueva_deuda = Deuda(
                    cliente_id=cliente_id,
                    tipo="Productos",
                    monto_inicial=costo_adicional,
                    saldo_pendiente=costo_adicional,
                    descripcion=linea_anexo,
                    estado="Pendiente"
                )
                db.session.add(nueva_deuda)

            db.session.commit()
            flash(f"✅ Se agregaron {cantidad} un. de {producto.nombre} a la cuenta de {cliente.nombre}.", "success")
        else:
            flash("❌ No hay suficiente stock en inventario.", "danger")
            
    return redirect(url_for('finanzas'))


@app.route('/abonar_deuda/<int:cliente_id>', methods=['POST'])
def abonar_deuda(cliente_id):
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))

    monto_abono = float(request.form.get('monto_abono', 0.0))
    moneda = request.form.get('moneda_abono')

    config = Configuracion.query.first()
    tasa = float(config.tasa_dolar) if config else 1.0
    monto_usd = monto_abono / tasa if moneda == 'BS' else monto_abono

    cliente = Cliente.query.get(cliente_id)
    if not cliente:
        flash("Cliente no encontrado.", "danger")
        return redirect(url_for('finanzas'))
    
    # Buscar la primera deuda pendiente que tenga el cliente para ir amortizando
    deuda = Deuda.query.filter_by(cliente_id=cliente_id, estado="Pendiente").first()
    
    if not deuda:
        flash("Este cliente no registra deudas activas individuales.", "warning")
        return redirect(url_for('finanzas'))

    # Registrar el historial analítico
    nuevo_abono = HistorialAbono(
        deuda_id=deuda.id,
        monto_usd=round(monto_usd, 2),
        monto_original=monto_abono,
        moneda_pago=moneda,
        tasa_momento=tasa
    )
    db.session.add(nuevo_abono)

    # Restar de los saldos
    deuda.saldo_pendiente -= monto_usd
    cliente.deuda_total -= monto_usd

    if deuda.saldo_pendiente <= 0.05: # Umbral de centavos por redondeo
        deuda.estado = "Pagada"
        deuda.saldo_pendiente = 0.0

    # =========================================================================
    # EFECTO REBOTAR EN CAJA CHICA: Creamos un Ingreso directo en tu mostrador
    # =========================================================================
    import time
    
    nuevo_ingreso_caja = Ingreso(
        tipo="Monto Externo",
        monto=round(monto_usd, 2),
        metodo_pago="Efectivo Dólar" if moneda == 'USD' else "Transferencia/Pago Móvil Bs",
        fecha=time.strftime('%Y-%m-%d'),  # <-- Genera "2026-07-14" de manera directa y segura
        cerrado=False
    )
    db.session.add(nuevo_ingreso_caja)
    db.session.commit()

    flash(f"Abono de ${monto_usd:.2f} USD procesado e integrado a la caja del día.", "success")
    return redirect(url_for('finanzas'))


@app.route('/liquidar_todo/<int:cliente_id>', methods=['POST'])
def liquidar_todo(cliente_id):
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))

    cliente = Cliente.query.get(cliente_id)
    if not cliente:
        flash("Cliente no encontrado.", "danger")
        return redirect(url_for('finanzas'))

    deudas_activas = Deuda.query.filter_by(cliente_id=cliente_id, estado="Pendiente").all()
    
    monto_recuperado = cliente.deuda_total

    for d in deudas_activas:
        d.estado = "Pagada"
        d.saldo_pendiente = 0.0
        
    cliente.deuda_total = 0.0

    # =========================================================================
    # EFECTO REBOTAR EN CAJA CHICA: Creamos un Ingreso directo en tu mostrador
    # =========================================================================
    import time

    # Inyectamos el dinero usando el modelo Ingreso con sus campos reales
    pago_completo_caja = Ingreso(
        tipo="Monto Externo", # O "Ingreso Liquidación" si prefieres mantener esa categoría
        monto=round(monto_recuperado, 2),
        metodo_pago="Efectivo Dólar",
        fecha=time.strftime('%Y-%m-%d'), # <-- Fecha segura sin variables intermedias
        cerrado=False
    )
    db.session.add(pago_completo_caja)
    db.session.commit()

    flash(f"Cuenta saldada por completo. Se ingresaron ${monto_recuperado:.2f} USD a caja.", "success")
    return redirect(url_for('finanzas'))

@app.route('/agregar_producto', methods=['POST'])
def agregar_producto():
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
        
    nombre = request.form.get('nombre')
    moneda_compra = request.form.get('moneda_compra')
    total_factura = float(request.form.get('total_factura', 0.0))
    stock_comprado = int(request.form.get('stock', 0))
    porcentaje_ganancia = float(request.form.get('porcentaje_ganancia', 0.0))
    
    # 1. MANEJO DEL IVA (16%)
    incluye_iva = request.form.get('incluye_iva') == 'true'
    if incluye_iva:
        total_factura = total_factura * 1.16
    
    # 2. CAPTURA DEL PROVEEDOR
    proveedor_input = request.form.get('proveedor_id')
    proveedor_id = int(proveedor_input) if proveedor_input and proveedor_input != "" else None
    
    # Tasas de cambio
    tasa_factura = float(request.form.get('tasa_factura', 1.0))
    config = Configuracion.query.first()
    tasa_negocio = float(config.tasa_dolar) if config else 1.0

    # A) COSTO REAL DEL PRODUCTO (Se calcula SIEMPRE sobre el total completo de la factura)
    if moneda_compra == 'BS':
        total_factura_usd = total_factura / tasa_factura
    else:
        total_factura_usd = total_factura

    costo_unitario_usd = (total_factura_usd / stock_comprado) if stock_comprado > 0 else 0.0
    precio_venta_usd = costo_unitario_usd * (1 + (porcentaje_ganancia / 100))

    # Guardamos el producto en inventario
    nuevo_producto = Producto(
        nombre=nombre,
        stock=stock_comprado,
        costo_compra=round(costo_unitario_usd, 2),
        precio_venta=round(precio_venta_usd, 2),
        proveedor_id=proveedor_id,
        ventas_totales=0  
    )
    db.session.add(nuevo_producto)

    # B) LOGICA DE FINANCIAMIENTO (Contado vs Crédito)
    tipo_pago = request.form.get('tipo_pago', 'contado')
    
    # Inicializamos la variable 'msg' aquí arriba para que SIEMPRE exista, vaya por donde vaya
    msg = f"Producto '{nombre}' registrado con éxito en el inventario al contado."
    
    if tipo_pago == 'credito':
        monto_inicial = float(request.form.get('monto_inicial', 0.0))
        monto_restante = float(request.form.get('monto_restante', 0.0))
        fecha_vencimiento = request.form.get('fecha_vencimiento')
        financista = request.form.get('financista') # 'Cashea', 'Proveedor', etc.

        # Si no se seleccionó fecha, la dejamos como None para la base de datos
        fecha_limite = fecha_vencimiento if fecha_vencimiento and fecha_vencimiento != "" else None

        # Convertimos la inicial a la tasa del negocio para el Gasto de caja
        if moneda_compra == 'BS':
            inicial_gasto_usd = monto_inicial / tasa_negocio
            restante_deuda_usd = monto_restante / tasa_factura
        else:
            inicial_gasto_usd = monto_inicial
            restante_deuda_usd = monto_restante

        # REGISTRO 1: El Gasto real en caja (Solo la Inicial)
        iva_txt = " + 16% IVA" if incluye_iva else ""
        descripcion_gasto = f"Inicial Compra Crédito ({financista}): {nombre}{iva_txt}"
        
        nuevo_gasto = Gasto(
            tipo="Proveedor",
            monto=round(inicial_gasto_usd, 2),
            descripcion=descripcion_gasto,
            metodo_pago="Efectivo Dólar" if moneda_compra == 'USD' else "Transferencia/Pago Móvil Bs",
            cerrado=False
        )
        db.session.add(nuevo_gasto)

        # Para que SQLite no proteste, el proveedor_id DEBE ser el del producto (no puede ser None)
        id_proveedor_seguro = proveedor_id if proveedor_id else 1 

        # Creamos una descripción bien detallada para saber que la deuda es con Cashea o el banco
        descripcion_deuda = f"Saldo restante compra {nombre}. Financiamiento vía: {financista}."

        # REGISTRO 2: La Deuda en Cuentas Por Pagar
        nueva_deuda = CuentaPorPagar(
            proveedor_id=id_proveedor_seguro,
            moneda=moneda_compra,
            monto_original=round(monto_restante, 2),
            tasa_factura=tasa_factura,
            monto=round(restante_deuda_usd, 2),
            descripcion=descripcion_deuda,
            fecha_limite=fecha_limite,
            pagado=False
        )
        db.session.add(nueva_deuda)
        
        # Redefinimos msg si fue crédito
        msg = f"Producto registrado. Inicial pagada a caja. Se generó la cuenta por pagar vinculada a {financista} por ${restante_deuda_usd:.2f} USD."

    else:
        # SI EL PAGO ES AL CONTADO: Registramos el egreso total de caja para que cuadren tus cuentas financieras
        iva_txt = " + 16% IVA" if incluye_iva else ""
        descripcion_gasto = f"Compra Contado: {nombre}{iva_txt}"
        
        # Convertimos el total de la factura a USD para la caja si vino en Bs
        if moneda_compra == 'BS':
            gasto_usd = total_factura / tasa_negocio
        else:
            gasto_usd = total_factura

        gasto_contado = Gasto(
            tipo="Proveedor",
            monto=round(gasto_usd, 2),
            descripcion=descripcion_gasto,
            metodo_pago="Efectivo Dólar" if moneda_compra == 'USD' else "Transferencia/Pago Móvil Bs",
            cerrado=False
        )
        db.session.add(gasto_contado)

    # Confirmamos todo en la base de datos de forma segura
    db.session.commit()
    flash(msg, "success")
    return redirect(url_for('finanzas'))

@app.route('/agregar_productos_lote', methods=['POST'])
def agregar_productos_lote():
    print("--- DATOS RECIBIDOS EN EL FORMULARIO ---")
    print("Nombres:", request.form.getlist('nombre[]'))
    print("Proveedor RAW:", request.form.get('proveedor_id'))
    print("Formulario completo:", request.form)
    print("---------------------------------------")

    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
        
    # 1. DATOS GENERALES DE LA FACTURA Y PROVEEDOR
    moneda_compra = request.form.get('moneda_compra', 'USD')
    tipo_pago = request.form.get('tipo_pago', 'contado')
    incluye_iva = request.form.get('incluye_iva') == 'true'
    
    # Procesamiento seguro del ID de proveedor
    proveedor_input = request.form.get('proveedor_id')
    proveedor_id = None
    if proveedor_input and proveedor_input.strip() and proveedor_input != "":
        try:
            proveedor_id = int(proveedor_input)
        except ValueError:
            proveedor_id = None

    tasa_factura = float(request.form.get('tasa_factura', 1.0))
    config = Configuracion.query.first()
    tasa_negocio = float(config.tasa_dolar) if config else 1.0

    # 2. CAPTURA DE LISTAS DINÁMICAS
    nombres = request.form.getlist('nombre[]')
    stocks = request.form.getlist('stock[]')
    totales_item = request.form.getlist('total_item[]')
    ganancias = request.form.getlist('porcentaje_ganancia[]')

    total_factura_usd = 0.0
    items_registrados = []

    # 3. PROCESAMIENTO Y FILTRADO ESTRICTO DE CADA RENGLÓN
    for i in range(len(nombres)):
        nombre_val = nombres[i]
        
        # Descarta nulos, vacíos o cadenas no válidas
        if not nombre_val or str(nombre_val).strip() == "" or str(nombre_val).strip() == "None":
            continue

        nombre_limpio = str(nombre_val).strip()

        # Parseo seguro de los campos numéricos
        try:
            stock_comprado = int(stocks[i]) if (i < len(stocks) and stocks[i]) else 0
            monto_item = float(totales_item[i]) if (i < len(totales_item) and totales_item[i]) else 0.0
            porcentaje_ganancia = float(ganancias[i]) if (i < len(ganancias) and ganancias[i]) else 0.0
        except (ValueError, TypeError):
            continue 

        if stock_comprado <= 0:
            continue

        # Aplicación de IVA si aplica
        if incluye_iva:
            monto_item = monto_item * 1.16

        # Conversión a USD
        if moneda_compra == 'BS':
            item_usd = monto_item / tasa_factura
        else:
            item_usd = monto_item

        total_factura_usd += item_usd
        costo_unitario_usd = (item_usd / stock_comprado) if stock_comprado > 0 else 0.0
        precio_venta_usd = costo_unitario_usd * (1 + (porcentaje_ganancia / 100))

        # Instanciar el producto con el proveedor asignado
        nuevo_producto = Producto(
            nombre=nombre_limpio,
            stock=stock_comprado,
            costo_compra=round(costo_unitario_usd, 2),
            precio_venta=round(precio_venta_usd, 2),
            proveedor_id=proveedor_id,
            ventas_totales=0  
        )
        db.session.add(nuevo_producto)
        items_registrados.append(nombre_limpio)

    # Si no hubo productos válidos, aborta
    if not items_registrados:
        flash("No se detectaron productos válidos con nombre y stock para guardar.", "warning")
        return redirect(url_for('finanzas'))

    # 4. MANEJO FINANCIERO (CONTADO VS CRÉDITO)
    resumen_nombres = ", ".join(items_registrados[:3]) + ("..." if len(items_registrados) > 3 else "")

    if tipo_pago == 'credito':
        monto_inicial = float(request.form.get('monto_inicial', 0.0))
        monto_restante = float(request.form.get('monto_restante', 0.0))
        fecha_vencimiento = request.form.get('fecha_vencimiento')
        financista = request.form.get('financista', 'Proveedor')

        fecha_limite = fecha_vencimiento if fecha_vencimiento and fecha_vencimiento != "" else None

        if moneda_compra == 'BS':
            inicial_gasto_usd = monto_inicial / tasa_negocio
            restante_deuda_usd = monto_restante / tasa_factura
        else:
            inicial_gasto_usd = monto_inicial
            restante_deuda_usd = monto_restante

        iva_txt = " + 16% IVA" if incluye_iva else ""
        
        nuevo_gasto = Gasto(
            tipo="Proveedor",
            monto=round(inicial_gasto_usd, 2),
            descripcion=f"Inicial Factura Lote ({financista}): {resumen_nombres}{iva_txt}",
            metodo_pago="Efectivo Dólar" if moneda_compra == 'USD' else "Transferencia/Pago Móvil Bs",
            cerrado=False
        )
        db.session.add(nuevo_gasto)

        id_proveedor_seguro = proveedor_id if proveedor_id else 1

        nueva_deuda = CuentaPorPagar(
            proveedor_id=id_proveedor_seguro,
            moneda=moneda_compra,
            monto_original=round(monto_restante, 2),
            tasa_factura=tasa_factura,
            monto=round(restante_deuda_usd, 2),
            descripcion=f"Restante Factura Lote: {resumen_nombres}. Vía: {financista}.",
            fecha_limite=fecha_limite,
            pagado=False
        )
        db.session.add(nueva_deuda)
        msg = f"Lote registrado ({len(items_registrados)} productos). Inicial deducida y deuda creada con {financista}."

    else:
        iva_txt = " + 16% IVA" if incluye_iva else ""
        gasto_contado = Gasto(
            tipo="Proveedor",
            monto=round(total_factura_usd, 2),
            descripcion=f"Compra Lote Contado: {resumen_nombres}{iva_txt}",
            metodo_pago="Efectivo Dólar" if moneda_compra == 'USD' else "Transferencia/Pago Móvil Bs",
            cerrado=False
        )
        db.session.add(gasto_contado)
        msg = f"Se registraron {len(items_registrados)} productos al contado con éxito."

    db.session.commit()
    flash(msg, "success")
    return redirect(url_for('finanzas'))

@app.route('/agregar_proveedor_finanzas', methods=['POST'])
def agregar_proveedor_finanzas():
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
        
    nombre = request.form.get('nombre')
    telefono = request.form.get('telefono')
    rubro = request.form.get('rubro') 

    if nombre:
        #  CORREGIDO: Eliminamos la línea 'deuda_pendiente=0.0' porque 
        # ahora las deudas se calculan solas de forma inteligente con la Fase 3.
        nuevo_proveedor = Proveedor(
            nombre=nombre,
            telefono=telefono,
            rubro=rubro if rubro else "General"
        )
        db.session.add(nuevo_proveedor)
        db.session.commit()
        flash(f"Proveedor '{nombre}' guardado correctamente.", "success")
    else:
        flash("El nombre del proveedor es obligatorio.", "danger")
        
    return redirect(url_for('finanzas'))

@app.route('/editar_producto/<int:producto_id>', methods=['POST'])
def editar_producto(producto_id):
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
        
    producto = Producto.query.get(producto_id)
    if not producto:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for('finanzas'))
        
    # Capturamos los campos permitidos incluyendo el nuevo stock
    nuevo_nombre = request.form.get('nombre')
    nuevo_stock = int(request.form.get('stock', 0))
    nuevo_precio = float(request.form.get('precio_venta', 0.0))
    
    # Aplicamos los cambios directo a la base de datos
    producto.nombre = nuevo_nombre
    producto.stock = nuevo_stock
    producto.precio_venta = round(nuevo_precio, 2)
    
    db.session.commit()
    flash(f"Producto '{producto.nombre}' actualizado con éxito (Stock actual: {producto.stock} unds).", "success")
    return redirect(url_for('finanzas'))


@app.route('/eliminar_producto/<int:producto_id>', methods=['POST'])
def eliminar_producto(producto_id):
    if not session.get('logueado'): 
        return redirect(url_for('bienvenida'))
        
    producto = Producto.query.get(producto_id)
    if not producto:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for('finanzas'))
        
    # SEGURIDAD REFORZADA: Doble verificación de stock cero en el backend
    if producto.stock > 0:
        flash("Seguridad: No se puede eliminar un producto que todavía tiene existencia en stock.", "danger")
        return redirect(url_for('finanzas'))
        
    db.session.delete(producto)
    db.session.commit()
    
    flash("Producto eliminado correctamente del inventario.", "success")
    return redirect(url_for('finanzas'))

def inicializar_base_de_datos():
    conn = sqlite3.connect('negocio.db') # Asegúrate de usar el mismo nombre de archivo .db que tienes en tus rutas
    cursor = conn.cursor()
    
    # Creamos la tabla usuarios con columnas para id, usuario y password
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        )
    ''')
    
    conn.commit()
    conn.close()

inicializar_base_de_datos()

@app.route('/activar-licencia', methods=['GET', 'POST'])
def activar_licencia():
    hwid = obtener_hwid()
    valida, mensaje, fecha = verificar_licencia_actual()
    config = Configuracion.query.first()

    if request.method == 'POST':
        nueva_key = request.form.get('licencia_key', '').strip().upper()
        
        # Guardamos temporalmente para verificar
        if not config:
            config = Configuracion(licencia_key=nueva_key)
            db.session.add(config)
        else:
            config.licencia_key = nueva_key
        
        db.session.commit()

        # Validamos si la clave recién guardada funciona
        es_valida, msg, fecha_venc = verificar_licencia_actual()
        if es_valida:
            if nueva_key.startswith("VITALICIA"):
                config.tipo_licencia = "Vitalicia"
                config.fecha_vencimiento_licencia = None
            else:
                config.tipo_licencia = "Temporal (Financiada/Mensual)"
                config.fecha_vencimiento_licencia = fecha_venc
            
            db.session.commit()
            flash("🎉 ¡Licencia activada con éxito!", "success")
            return redirect(url_for('bienvenida'))
        else:
            flash(f"❌ Error: {msg}", "danger")

    return render_template('activar_licencia.html', hwid=hwid, valida=valida, mensaje=mensaje, config=config)

if __name__ == '__main__':
    
    with app.app_context():
        db.create_all()  
        
    app.run(debug=True)